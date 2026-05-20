from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from flask_cors import CORS
import os
import json
import logging
import io
import pymongo
from bson import ObjectId
from fpdf import FPDF
import mysql.connector
from mysql.connector import pooling
from contextlib import contextmanager
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import secrets
from PIL import Image
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Border, Side
from Final_test import Generate_test_timetable
from dotenv import load_dotenv
load_dotenv()
app = Flask(__name__, template_folder="templates", static_folder='static')
CORS(app)
app.secret_key = os.urandom(24)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Configure logging
app.logger.setLevel(logging.INFO)

# ── MySQL Connection Pool ────────────────────────────────────────────────────
# Pool size 20 — supports concurrent requests without exhaustion.
# pool_reset_session=True ensures each connection is clean on checkout.
_POOL_SIZE = int(os.environ.get("DB_POOL_SIZE", "20"))
_mysql_pool = pooling.MySQLConnectionPool(
    pool_name="timetable_pool",
    pool_size=_POOL_SIZE,
    pool_reset_session=True,
    host=os.environ.get("DB_HOST"),
    user=os.environ.get("DB_USER"),
    password=os.environ.get("DB_PASSWORD"),
    database=os.environ.get("DB_NAME"),
    connection_timeout=10,
    auth_plugin='mysql_native_password',
)


def get_mysql_connection():
    """Return a raw pooled connection (caller must close it — prefer mysql_connection())."""
    return _mysql_pool.get_connection()


@contextmanager
def mysql_connection():
    """Context manager that guarantees the pooled connection is returned on exit.

    Usage:
        with mysql_connection() as (cnx, cur):
            cur.execute(...)
            rows = cur.fetchall()
    """
    cnx = _mysql_pool.get_connection()
    cur = cnx.cursor(dictionary=True)
    try:
        yield cnx, cur
    except Exception:
        try:
            cnx.rollback()
        except Exception:
            pass
        raise
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            cnx.close()  # returns connection to pool
        except Exception:
            pass

# MongoDB Connection — cached at module level to avoid reconnecting per request
_mongo_client = pymongo.MongoClient(
    os.environ.get("MONGODB", "mongodb://localhost:27017/")
)

def get_mongodb_connection():
    return _mongo_client["timetablegenius"]

# Login Required Decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Roles Required Decorator
def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if session.get('user_role') not in roles:
                return jsonify({'error': 'Unauthorized. Admin role required.'}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_effective_user_id():
    """Returns the admin user_id if the current user is a staff/teacher, otherwise returns current user_id."""
    return session.get('parent_admin_id') or session.get('user_id')

def parse_time_robustly(time_str):
    """Parses time strings like '8am', '8:00am', '08:00am' robustly."""
    time_str = time_str.lower().strip()
    formats = ['%I:%M%p', '%I%p', '%H:%M', '%H']
    for fmt in formats:
        try:
            return datetime.strptime(time_str, fmt)
        except ValueError:
            continue
    # Fallback to a very simple parse if nothing else works
    try:
        if 'am' in time_str or 'pm' in time_str:
            parts = time_str.replace('am', '').replace('pm', '').split(':')
            hour = int(parts[0])
            minute = int(parts[1]) if len(parts) > 1 else 0
            if 'pm' in time_str and hour < 12: hour += 12
            if 'am' in time_str and hour == 12: hour = 0
            return datetime(1900, 1, 1, hour, minute)
    except Exception:
        pass
    return datetime(1900, 1, 1)

class TimetablePDF(FPDF):
    def header(self):
        # Logo
        try:
            logo_path = os.path.join(app.static_folder, 'icon.jpg')
            if os.path.exists(logo_path):
                self.image(logo_path, 10, 8, 33)
        except Exception:
            pass
        # Arial bold 15
        self.set_font('Helvetica', 'B', 15)
        # Move to the right
        self.cell(80)
        # Title
        # Fix for deprecation ln=0 (use new_x and new_y)
        self.cell(30, 10, 'Timetable Genius', border=0, align='C')
        # Line break
        self.ln(20)

    def footer(self):
        # Position at 1.5 cm from bottom
        self.set_y(-15)
        # Arial italic 8
        self.set_font('Helvetica', 'I', 8)
        # Page number
        self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', border=0, align='C')

@app.route('/api/export/pdf', methods=['GET'])
@login_required
def export_pdf():
    eff_user_id = get_effective_user_id()
    timetable_id = request.args.get('id')
    category_type = request.args.get('category_type', 'class')
    filter_value = request.args.get('filter_value')

    try:
        mongo_db = get_mongodb_connection()
        if not timetable_id or timetable_id == 'undefined':
            timetable_doc = mongo_db["timetables"].find_one({"user_id": user_id}, sort=[("generated_at", -1)])
        else:
            timetable_doc = mongo_db["timetables"].find_one({"_id": ObjectId(timetable_id), "user_id": user_id})
        
        if not timetable_doc:
            return jsonify({'error': 'Timetable not found'}), 404
        
        generated_timetable = timetable_doc['generated_timetable']
        if isinstance(generated_timetable, str):
            generated_timetable = json.loads(generated_timetable)
        
        pdf = TimetablePDF(orientation='L', unit='mm', format='A4')
        pdf.alias_nb_pages()
        
        # Define the classes/entities to include
        if category_type == 'class':
            if filter_value and filter_value.lower() != 'all':
                target_class = next((k for k in generated_timetable.keys() if k.lower() == filter_value.lower()), None)
                entities = [target_class] if target_class else []
            else:
                entities = list(generated_timetable.keys())
        elif category_type == 'teacher':
            entities = [filter_value]
        elif category_type == 'classroom':
            entities = [filter_value]
        else:
            entities = []

        for entity_name in entities:
            if not entity_name: continue
            
            pdf.add_page()
            pdf.set_font('Helvetica', 'B', 12)
            # Fix for deprecation ln=1 (use new_x and new_y implied by ln() or specific parameters)
            pdf.cell(0, 10, f"{category_type.capitalize()}: {entity_name}", border=0, align='C', new_x="LMARGIN", new_y="NEXT")
            pdf.ln(5)

            # Get schedule data
            if category_type == 'class':
                schedule_data = generated_timetable.get(entity_name)
            elif category_type == 'teacher':
                schedule_data = filter_timetable_by(generated_timetable, entity_name, 'teacher')
            elif category_type == 'classroom':
                schedule_data = filter_timetable_by(generated_timetable, entity_name, 'classroom')
            
            if not schedule_data:
                pdf.cell(0, 10, "No data available", border=0, new_x="LMARGIN", new_y="NEXT")
                continue

            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
            all_slots = set()
            for day in days:
                if day in schedule_data:
                    all_slots.update(schedule_data[day].keys())
            
            # Use robust time parsing for sorting
            sorted_slots = sorted(list(all_slots), key=lambda x: parse_time_robustly(x.split(' to ')[0]))

            # Table Header
            pdf.set_font('Helvetica', 'B', 10)
            col_width = (pdf.w - 20) / 6.0
            pdf.cell(col_width, 10, 'Time', border=1, align='C')
            for day in days:
                pdf.cell(col_width, 10, day, border=1, align='C')
            pdf.ln()

            # Table Rows
            pdf.set_font('Helvetica', '', 8)
            for slot in sorted_slots:
                # Calculate required height for this row
                max_lines = 1
                row_data = []
                for day in days:
                    cell = schedule_data.get(day, {}).get(slot, '-')
                    if isinstance(cell, dict):
                        s_type = cell.get('type', 'Theory')
                        content = f"{cell['subject']}\n({cell.get('classroom', cell.get('className', ''))})\n[{s_type}]"
                        row_data.append({'content': content, 'type': s_type})
                    else:
                        content = str(cell)
                        row_data.append({'content': content, 'type': None})
                    max_lines = max(max_lines, content.count('\n') + 1)
                
                row_height = max_lines * 5
                
                # Check for page break
                if pdf.get_y() + row_height > pdf.page_break_trigger:
                    pdf.add_page()
                    # Re-draw header
                    pdf.set_font('Helvetica', 'B', 10)
                    pdf.cell(col_width, 10, 'Time', border=1, align='C')
                    for day in days:
                        pdf.cell(col_width, 10, day, border=1, align='C')
                    pdf.ln()
                    pdf.set_font('Helvetica', '', 8)

                # Draw cells
                x_start = pdf.get_x()
                y_start = pdf.get_y()
                pdf.multi_cell(col_width, 5, slot, border=1, align='C')
                pdf.set_xy(x_start + col_width, y_start)
                
                for item in row_data:
                    curr_x = pdf.get_x()
                    
                    # Design based on type
                    if item['type'] == 'Theory':
                        pdf.set_fill_color(230, 240, 255) # Light Blue
                        pdf.multi_cell(col_width, 5, item['content'], border=1, align='C', fill=True)
                    elif item['type'] == 'Practical':
                        pdf.set_fill_color(230, 255, 230) # Light Green
                        pdf.multi_cell(col_width, 5, item['content'], border=1, align='C', fill=True)
                    else:
                        pdf.multi_cell(col_width, 5, item['content'], border=1, align='C', fill=False)
                        
                    pdf.set_xy(curr_x + col_width, y_start)
                pdf.ln(row_height)

            # Add Total Weekly Hours at the end of each entity's schedule
            total_lectures = sum(1 for day in days for slot in sorted_slots if isinstance(schedule_data.get(day, {}).get(slot), dict))
            lecture_params = timetable_doc.get('parameters', {})
            lecture_duration = int(lecture_params.get('lecture_duration', 45) if lecture_params else 45)
            total_hours = (total_lectures * lecture_duration) / 60.0
            
            pdf.ln(5)
            pdf.set_font('Helvetica', 'B', 10)
            pdf.cell(0, 10, f"Total Weekly Hours: {total_hours:.2f} hrs", border=0, align='R', new_x="LMARGIN", new_y="NEXT")

        pdf_out = pdf.output(dest='S')
        return send_file(io.BytesIO(pdf_out), as_attachment=True, download_name=f"timetable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf", mimetype='application/pdf')

    except Exception as e:
        app.logger.error(f"PDF Export error: {e}")
        return jsonify({'error': str(e)}), 500

# Routes
@app.route('/')
def index():
    # Already logged in → go straight to dashboard
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('landing.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Already logged in → skip login page
    if request.method == 'GET' and 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        data = request.get_json()
        
        email = data.get('email')
        password = data.get('password')
        
        try:
            with mysql_connection() as (mysql, cursor):
                cursor.execute("""
                    SELECT user_id, name, email, password_hash, role, parent_admin_id
                    FROM users 
                    WHERE email = %s
                """, (email,))
            
                user = cursor.fetchone()
            
                if user and check_password_hash(user['password_hash'], password):
                    session['user_id'] = user['user_id']
                    session['user_name'] = user['name']
                    session['user_role'] = user['role']
                    session['parent_admin_id'] = user['parent_admin_id']
                    return jsonify({'redirect': '/dashboard'})
                else:
                    return jsonify({'error': 'Invalid credentials'}), 401
        except Exception as e:
            print(f"Error during login: {e}")
            return jsonify({'error': 'Login failed'}), 500
    return render_template('Sign-in.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        data = request.get_json()
        
        name = data.get('name')
        email = data.get('email')
        password = data.get('password')
        role = data.get('role')
        parent_admin_id = data.get('parent_admin_id') # For staff creation
        
        if not all([name, email, password, role]):
            return jsonify({'error': 'All fields are required'}), 400
        
        # Hash the password
        hashed_password = generate_password_hash(password)
        
        try:
            with mysql_connection() as (mysql, cur):
                # Check if user already exists
                cur.execute("SELECT email FROM users WHERE email = %s", (email,))
                existing_user = cur.fetchone()
            
                if existing_user:
                    return jsonify({'error': 'Email already exists'}), 409
            
                # Insert new user
                cur.execute("""
                    INSERT INTO users (name, email, password_hash, role, parent_admin_id, created_at) 
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (name, email, hashed_password, role, parent_admin_id, datetime.now()))
            
                mysql.commit()
                return jsonify({'message': 'Registration successful'}), 201
            
        except Exception as e:
            print(f"Error during signup: {e}")
            return jsonify({'error': 'Registration failed'}), 500
    return render_template('Sign-up.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({'error': 'Email is required'}), 400
            
        
        try:
            with mysql_connection() as (mysql_conn, cursor):
                cursor.execute("SELECT user_id, name FROM users WHERE email = %s", (email,))
                user = cursor.fetchone()
            
                if user:
                    # Generate token
                    token = secrets.token_urlsafe(32)
                    expiry = datetime.now() + timedelta(hours=1)
                
                    # Store token in DB (Assuming columns reset_token and token_expiry exist)
                    # If they don't, this will fail and we'll handle it
                    try:
                        cursor.execute("""
                            UPDATE users 
                            SET reset_token = %s, token_expiry = %s 
                            WHERE user_id = %s
                        """, (token, expiry, user['user_id']))
                        mysql_conn.commit()
                    
                        # Log the reset link (since we don't have real mail setup)
                        reset_link = url_for('reset_password', token=token, _external=True)
                        print(f"\n[PASSWORD RESET] For user {user['name']} ({email}):")
                        print(f"Link: {reset_link}\n")
                    
                        return jsonify({'message': 'Reset instructions sent to your email'})
                    except Exception as db_err:
                        app.logger.error(f"Database error during token storage: {db_err}")
                        return jsonify({'error': 'Database schema might need updates for reset tokens'}), 500
            
                # For security, always return success message even if email doesn't exist
                return jsonify({'message': 'If an account exists with that email, reset instructions have been sent'})
            
        except Exception as e:
            print(f"Error in forgot_password: {e}")
            return jsonify({'error': 'An error occurred'}), 500
    return render_template('forgot-password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if request.method == 'POST':
        data = request.get_json()
        new_password = data.get('password')
        
        if not new_password:
            return jsonify({'error': 'Password is required'}), 400
            
        
        try:
            with mysql_connection() as (mysql_conn, cursor):
                # Check token and expiry
                cursor.execute("""
                    SELECT user_id FROM users 
                    WHERE reset_token = %s AND token_expiry > %s
                """, (token, datetime.now()))
            
                user = cursor.fetchone()
            
                if not user:
                    return jsonify({'error': 'Invalid or expired token'}), 400
                
                # Update password
                hashed_password = generate_password_hash(new_password)
                cursor.execute("""
                    UPDATE users 
                    SET password_hash = %s, reset_token = NULL, token_expiry = NULL 
                    WHERE user_id = %s
                """, (hashed_password, user['user_id']))
                mysql_conn.commit()
            
                return jsonify({'message': 'Password reset successful'})
            
        except Exception as e:
            print(f"Error in reset_password: {e}")
            return jsonify({'error': 'An error occurred'}), 500
    return render_template('reset-password.html', token=token)

@app.route('/dashboard')
@login_required
def dashboard():
    
    try:
        eff_user_id = get_effective_user_id()
        with mysql_connection() as (mysql, cur):
            # Get user details first
            cur.execute(
                "SELECT * FROM users WHERE user_id = %s",
                (session.get('user_id'),)
            )
            user_details = cur.fetchone()
        
            if not user_details:
                return redirect(url_for('login'))

            # Get counts for dashboard
            cur.execute("SELECT COUNT(*) as count FROM teachers WHERE user_id = %s", (eff_user_id,))
            teacher_count = cur.fetchone()['count']
        
            cur.execute("SELECT COUNT(*) as count FROM classes WHERE user_id = %s", (eff_user_id,))
            class_count = cur.fetchone()['count']
        
            cur.execute("SELECT COUNT(*) as count FROM classrooms WHERE user_id = %s", (eff_user_id,))
            classroom_count = cur.fetchone()['count']
        
            cur.execute("SELECT COUNT(*) as count FROM subjects WHERE user_id = %s", (eff_user_id,))
            subject_count = cur.fetchone()['count']
        
            cur.execute("""
                SELECT a.*, u.name as user_name 
                FROM activity a
                LEFT JOIN users u ON a.user_id = u.user_id
                WHERE a.user_id = %s
                ORDER BY a.timestamp DESC 
                LIMIT 5
            """, (eff_user_id,))
            activities = cur.fetchall()

            # Return JSON if requested
            if request.headers.get('Content-Type') == 'application/json':
                return jsonify({
                    'user_name': user_details['name'],
                    'stats': {
                        'teachers': teacher_count,
                        'classes': class_count,
                        'classrooms': classroom_count,
                        'subjects': subject_count
                    },
                    'activities': activities
                })

            # Otherwise return the rendered template
            return render_template('dashboard.html',
                user_name=user_details['name'],
                stats={
                    'teachers': teacher_count,
                    'classes': class_count,
                    'classrooms': classroom_count,
                    'subjects': subject_count
                },
                activities=activities
            )
        
    except Exception as e:
        print(f"Error loading dashboard: {e}")
        return render_template('dashboard.html', error="Failed to load dashboard data")
@app.route('/api/teachers', methods=['GET'])
@login_required
def get_teachers():
    try:
        eff_user_id = get_effective_user_id()
        with mysql_connection() as (mysql, cur):
            # Fixed SQL query syntax (removed extra comma, added missing column
            cur.execute("""
                        SELECT 
                        t.teacher_id, 
                        t.teacher_name, 
                        t.email, 
                        t.phone, 
                        t.weekly_hours, 
                        d.name AS department_name, 
                        GROUP_CONCAT(CONCAT(s.name, '-', SUBSTRING(s.type, 1, 1)) SEPARATOR ', ') AS subjects
                    FROM 
                        teachers t
                    JOIN 
                        departments d ON t.department_id = d.department_id
                    LEFT JOIN 
                        teacher_subjects ts ON t.teacher_id = ts.teacher_id
                    LEFT JOIN 
                        subjects s ON ts.subject_id = s.subject_id
                    WHERE 
                        t.user_id = %s
                    GROUP BY 
                        t.teacher_id, t.teacher_name, t.email, t.phone, t.weekly_hours, d.name;
                        """, (eff_user_id,))
            teachers = cur.fetchall()
            return jsonify(teachers)
    except Exception as e:
        app.logger.error(f"Error fetching teachers: {e}")
        return jsonify({'error': str(e)}), 500
@app.route('/api/teachers', methods=['POST'])
@login_required
@roles_required('admin')
def add_teacher():
    teacher = request.json
    
    try:
        with mysql_connection() as (mysql, cur):
            # Get or create department
            cur.execute("SELECT department_id FROM departments WHERE name = %s", (teacher.get('department'),))
            department = cur.fetchone()
        
            if not department:
                # Create new department
                cur.execute("INSERT INTO departments (name) VALUES (%s)", (teacher.get('department'),))
                mysql.commit()
                department_id = cur.lastrowid
            else:
                department_id = department['department_id']
        
            # Insert teacher
            cur.execute("""
                INSERT INTO teachers (user_id, department_id, teacher_name, weekly_hours, email, phone) 
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                session.get('user_id'),
                department_id,
                teacher.get('name'),
                teacher.get('weeklyhour'),
                teacher.get('email'),
                teacher.get('phone')
            ))

                
            # Log activity
            cur.execute("""
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id) 
                VALUES (%s, %s, %s, %s, %s)
            """, (
                session.get('user_id'),
                'CREATE',
                f'Added teacher {teacher.get("name")}',
                'teacher',
                cur.lastrowid
            ))
            mysql.commit()
        
            return jsonify({'message': 'Teacher added successfully', 'teacher_id': cur.lastrowid}), 201
    
    except Exception as e:
        mysql.rollback()
        print(f"Error adding teacher: {e}")
        return jsonify({'error': 'Failed to add teacher: ' + str(e)}), 500
@app.route('/api/teachers/<int:teacher_id>', methods=['GET'])
@login_required
def get_teacher(teacher_id):
    try:
        # Get teacher data from database
        mysql=get_mysql_connection()
        cur = mysql.cursor(dictionary=True)
        cur.execute("""
                    SELECT t.user_id, t.department_id, t.teacher_name, t.weekly_hours,
                    t.email, t.phone, d.name as department_name
                    FROM teachers t
                    JOIN departments d ON t.department_id = d.department_id
                    WHERE t.teacher_id = %s AND t.user_id = %s
                    """, (teacher_id, session.get('user_id')))

        teacher = cur.fetchone()
        if not teacher:
            return jsonify({'error': 'Teacher not found'}), 404
        
        if request.headers.get('Content-Type') == 'application/json':
            return jsonify(teacher)        

        
        
    except Exception as e:
        print(f"Error fetching teacher: {e}")
        return jsonify({'error': 'Failed to fetch teacher'}), 500
    finally:
        cur.close()
        mysql.close()

@app.route('/api/teachers/<int:teacher_id>', methods=['PUT'])
@login_required
@roles_required('admin')
def update_teacher(teacher_id):
    teacher = request.json
    
    try:
        with mysql_connection() as (mysql, cur):
            # Get or create department
            cur.execute("SELECT department_id FROM departments WHERE name = %s", (teacher.get('department'),))
            department = cur.fetchone()
        
            if not department:
                # Create new department
                cur.execute("INSERT INTO departments (name) VALUES (%s)", (teacher.get('department'),))
                mysql.commit()
                department_id = cur.lastrowid
            else:
                department_id = department['department_id']
        
            # Update teacher
            cur.execute("""
                UPDATE teachers 
                SET department_id = %s, teacher_name = %s, weekly_hours = %s, email = %s, phone = %s
                WHERE teacher_id = %s AND user_id = %s
            """, (
                department_id,
                teacher.get('name'),
                teacher.get('weekly_hours'),
                teacher.get('email'),
                teacher.get('phone'),
                teacher_id,
                session.get('user_id')
            ))
        
            if cur.rowcount == 0:
                return jsonify({'error': 'Teacher not found or not authorized'}), 404
        
            mysql.commit()
        
            # Log activity
            cur.execute("""
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id) 
                VALUES (%s, %s, %s, %s, %s)
            """, (
                session.get('user_id'),
                'UPDATE',
                f'Updated teacher {teacher.get("name")}',
                'teacher',
                teacher_id
            ))
            mysql.commit()
        
            return jsonify({'message': 'Teacher updated successfully'})
    
    except Exception as e:
        mysql.rollback()
        print(f"Error updating teacher: {e}")
        return jsonify({'error': 'Failed to update teacher: ' + str(e)}), 500
@app.route('/api/teachers/<int:teacher_id>', methods=['DELETE'])
@login_required
@roles_required('admin')
def delete_teacher(teacher_id):
    try:
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                SELECT * FROM teachers 
                WHERE teacher_id = %s AND user_id = %s
            """, (teacher_id, session.get('user_id')))
            teacher = cur.fetchone()
            if teacher is None:
                return jsonify({'error': 'Teacher not found or not authorized'}), 404

            teacher_name = teacher['teacher_name']
            
            # Delete teacher
            cur.execute("""
                DELETE FROM teachers WHERE teacher_id = %s AND user_id = %s
            """, (teacher_id, session.get('user_id')))
            mysql.commit()
            
            # Log activity
            cur.execute("""
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                session.get('user_id'),
                'DELETE',
                f'Deleted teacher {teacher_name}',
                'teacher',
                teacher_id
            ))
            mysql.commit()
            return jsonify({'message': 'Teacher deleted successfully'}), 200
    except Exception as e:
        app.logger.error(f"Error deleting teacher: {e}")
        return jsonify({'error': 'Failed to delete teacher: ' + str(e)}), 500

@app.route('/api/teachers/<int:teacher_id>/availability', methods=['GET'])
@login_required
def get_teacher_availability(teacher_id):
    try:
        with mysql_connection() as (mysql, cur):
            # Verify teacher belongs to user
            cur.execute("SELECT teacher_id FROM teachers WHERE teacher_id = %s AND user_id = %s", (teacher_id, session.get('user_id')))
            if not cur.fetchone():
                return jsonify({'error': 'Unauthorized'}), 403
            
            cur.execute("""
                SELECT availability_id, day_of_week, 
                       TIME_FORMAT(start_time, '%%H:%%i') as start_time, 
                       TIME_FORMAT(end_time, '%%H:%%i') as end_time, 
                       preference_level
                FROM teacher_availability 
                WHERE teacher_id = %s
                ORDER BY FIELD(day_of_week, 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'), start_time
            """, (teacher_id,))
            availability = cur.fetchall()
            return jsonify(availability)
    except Exception as e:
        app.logger.error(f"Error fetching availability: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/teachers/<int:teacher_id>/availability', methods=['POST'])
@login_required
@roles_required('admin')
def update_teacher_availability(teacher_id):
    data = request.json
    try:
        with mysql_connection() as (mysql, cur):
            # Verify teacher belongs to user
            cur.execute("SELECT teacher_id FROM teachers WHERE teacher_id = %s AND user_id = %s", (teacher_id, session.get('user_id')))
            if not cur.fetchone():
                return jsonify({'error': 'Unauthorized'}), 403
            
            # Simple approach: Clear and re-add
            cur.execute("DELETE FROM teacher_availability WHERE teacher_id = %s", (teacher_id,))
            
            for slot in data.get('availability', []):
                cur.execute("""
                    INSERT INTO teacher_availability (teacher_id, day_of_week, start_time, end_time, preference_level)
                    VALUES (%s, %s, %s, %s, %s)
                """, (teacher_id, slot['day_of_week'], slot['start_time'], slot['end_time'], slot['preference_level']))
            
            mysql.commit()
            
            # Log activity
            cur.execute("INSERT INTO activity (user_id, action_type, description, entity_type, entity_id) VALUES (%s, %s, %s, %s, %s)",
                       (session.get('user_id'), 'UPDATE', f"Updated availability for teacher {teacher_id}", 'teacher', teacher_id))
            mysql.commit()
            
            return jsonify({'success': True})
    except Exception as e:
        app.logger.error(f"Error updating availability: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/classes', methods=['GET'])
@login_required
def get_classes():
    try:
        eff_user_id = get_effective_user_id()
        with mysql_connection() as (mysql, cur):
            cur.execute("""
            SELECT 
                c.class_id, 
                c.name AS class_name, 
                g.name AS grade_name, 
                cr.room_number, 
                c.students_count,
                GROUP_CONCAT(CONCAT(s.name, '-', SUBSTRING(s.type, 1, 1)) SEPARATOR ', ') AS subjects
            FROM classes c
            LEFT JOIN grades g ON c.grade_id = g.grade_id
            LEFT JOIN classrooms cr ON c.classroom_id = cr.classroom_id
            LEFT JOIN class_subjects cs ON c.class_id = cs.class_id
            LEFT JOIN subjects s ON cs.subject_id = s.subject_id
            WHERE c.user_id = %s
            GROUP BY c.class_id, c.name, g.name, cr.room_number, c.students_count
            """, (eff_user_id,))
            classes = cur.fetchall()
            return jsonify(classes)
    except Exception as e:
        print(f"Error fetching classes: {e}")
        return jsonify({'error': str(e)}), 500
@app.route('/api/classes/<int:classes_id>',methods=['GET'])
@login_required
def get_class(classes_id):
    mysql = get_mysql_connection()
    cur=mysql.cursor(
        dictionary=True
    )
    
    try:
        cur.execute("""
                    SELECT
                    c.class_id,
                    c.name AS class_name,
                    g.name AS grade_name,
                    cr.room_number,
                    c.students_count,
                    c.classroom_id
                    FROM classes c
                    LEFT JOIN grades g ON c.grade_id = g.grade_id
                    LEFT JOIN classrooms cr ON c.classroom_id = cr.classroom_id
                    WHERE c.class_id = %s
                    """, (classes_id,))
        class_info = cur.fetchone()
        if class_info is None:
            return jsonify({'error': 'Class not found'}), 404
        
        cur.execute("""
                    SELECT
                    s.subject_id,
                    CONCAT(s.name, '-', SUBSTRING(s.type, 1, 1)) AS subjects
                    FROM class_subjects cs
                    JOIN subjects s ON cs.subject_id = s.subject_id
                    WHERE cs.class_id = %s
                    """, (classes_id,))
        
        subjects=cur.fetchall()
        return jsonify({'class_info': class_info, 'subjects': subjects})
    except Exception as e:
        print(f"Error fetching class: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        mysql.close()
         
@app.route('/api/classes', methods=['POST'])
@login_required
@roles_required('admin')
def add_class():
    class_data = request.json
    
    try:
        with mysql_connection() as (mysql, cur):
            # Start transaction
            mysql.start_transaction()
        
            # Check if grade exists
            cur.execute("SELECT grade_id FROM grades WHERE name = %s", (class_data.get('grade'),))
            grade = cur.fetchone()
        
            if not grade:
                # Create new grade if it doesn't exist
                cur.execute("INSERT INTO grades (name) VALUES (%s)", (class_data.get('grade'),))
                grade_id = cur.lastrowid
            else:
                grade_id = grade['grade_id']
        
            # Insert class
            cur.execute("""
                INSERT INTO classes (user_id, name, grade_id, classroom_id, students_count) 
                VALUES (%s, %s, %s, %s, %s)
            """, (
                session.get('user_id'),
                class_data.get('name'),
                grade_id,
                class_data.get('classroom') if class_data.get('classroom') else None,  # Handle empty classroom selection
                int(class_data.get('student', 0))  # Ensure integer conversion for student count
            ))
        
            # Get the newly created class_id
            class_id = cur.lastrowid
        
            # Add subjects to class_subjects table
            if 'subjects' in class_data and class_data['subjects']:
                subjects = class_data['subjects']
                if isinstance(subjects, str):
                    # If subjects is a comma-separated string, split it
                    subjects = subjects.split(',')
            
                for subject_id in subjects:
                    # Skip empty strings
                    if not subject_id:
                        continue
                    
                    cur.execute("""
                        INSERT INTO class_subjects (class_id, subject_id) 
                        VALUES (%s, %s)
                    """, (class_id, int(subject_id)))
        
            # Log activity
            cur.execute("""
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id) 
                VALUES (%s, %s, %s, %s, %s)
            """, (
                session.get('user_id'),
                'CREATE',
                f'Added class {class_data.get("name")}',
                'class',
                class_id
            ))
        
            # Commit all changes
            mysql.commit()
        
            return jsonify({
                'success': True,
                'message': 'Class added successfully', 
                'class_id': class_id
            }), 201
    
    except Exception as e:
        # Rollback transaction on error
        mysql.rollback()
        print(f"Error adding class: {e}")
        return jsonify({
            'success': False,
            'error': f'Failed to add class: {str(e)}'
        }), 500
    
        # Close database connections
@app.route('/api/classes/<int:class_id>', methods=['PUT'])
@login_required
@roles_required('admin')
def update_class(class_id):
    # Get class data from request
    class_data = request.get_json()
    
    try:
        with mysql_connection() as (mysql, cur):
            # Get class data from request
            class_data = request.get_json()
        
            if not class_data:
                return jsonify({"error": "No data provided"}), 400
        
            # Extract values from the request data
            name = class_data.get('name')
            grade_id = None
            classroom_id = class_data.get('classroom')
            students_count = int(class_data.get('student', 0))
            subjects=class_data.get('subjects',[])
        
            # Get grade_id from grade name
            if 'grade' in class_data and class_data['grade']:
                cur.execute("SELECT grade_id FROM grades WHERE name = %s", (class_data['grade'],))
                grade_result = cur.fetchone()
                if grade_result:
                    grade_id = grade_result['grade_id']
                else:
                    cur.execute("INSERT INTO grades(name) VALUES(%s)",(
                        class_data['grade'],
                    ))
                    mysql.commit()
                    grade_id=cur.lastrowid
        
            # Update class information
            update_query = """
                UPDATE classes 
                SET name = %s, grade_id = %s, classroom_id = %s, students_count = %s            
                WHERE class_id = %s AND user_id=%s
            """
        
            cur.execute(update_query, (
                name, 
                grade_id, 
                classroom_id if classroom_id else None, 
                students_count, 
                class_id,
                session.get('user_id'),
            ))
        
            # Delete existing subject relationships
            cur.execute("DELETE FROM class_subjects WHERE class_id = %s", (class_id,))
            mysql.commit()
            # Insert new subject relationships
            if subjects and len(subjects) > 0:
                for subject_id in subjects:
                    if subject_id:  
                        cur.execute(
                            "INSERT INTO class_subjects (class_id, subject_id) VALUES (%s, %s)",
                            (class_id, int(subject_id),)
                        )
        
            # Commit changes
            mysql.commit()
        
            # Log the activity
            cur.execute(
                "INSERT INTO activity (user_id, action_type, description, entity_type, entity_id) VALUES (%s, %s, %s, %s, %s)",
                (session.get("user_id"), "UPDATE", f"Updated class: {name}", "class", class_id,)
            )
            mysql.commit()
        
            return jsonify({"success": True, "message": "Class updated successfully"})
        
    except Exception as e:
        # Rollback in case of error
        mysql.rollback()
        return jsonify({"error": str(e)}), 500
@app.route('/api/classes/<int:class_id>', methods=['DELETE'])
@login_required
@roles_required('admin')
def delete_class(class_id):

    try:
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                SELECT * FROM classes
                WHERE class_id = %s AND user_id = %s
            """, (class_id, session.get('user_id')))
            classs = cur.fetchone()
            if classs is None:
                return jsonify({"error": "Class not found"}), 404

            cur.execute(
                "DELETE FROM class_subjects WHERE class_id = %s",
                (class_id,),
            )
            cur.execute(
                "DELETE FROM classes WHERE class_id = %s AND user_id = %s",
                (class_id, session.get("user_id")),
            )

            # Log activity
            cur.execute(
                "INSERT INTO activity (user_id, action_type, description, entity_type, entity_id) VALUES (%s, %s, %s, %s, %s)",
                (session.get("user_id"), "DELETE", f"Deleted class: {classs['name']}", "class", class_id),
            )
            mysql.commit()
            return jsonify({"message": "Class deleted successfully"}), 200
    except Exception as e:
        app.logger.error(f"Error deleting class: {e}")
        mysql.rollback()
        return jsonify({"error": str(e)}), 500
@app.route('/api/subjects', methods=['GET'])
@login_required
def get_subjects():
    try:
        eff_user_id = get_effective_user_id()
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                SELECT s.subject_id, s.name, s.type, s.weekly_hours, d.name as department_name, t.teacher_name
                FROM subjects s
                JOIN departments d ON s.department_id = d.department_id
                JOIN teacher_subjects ts ON s.subject_id = ts.subject_id
                JOIN teachers t ON ts.teacher_id = t.teacher_id
                WHERE s.user_id = %s
            """, (eff_user_id,))
            subjects = cur.fetchall()
            return jsonify(subjects)
    except Exception as e:
        print(f"Error fetching subjects: {e}")
        return jsonify({'error': str(e)}), 500
@app.route('/api/subjects', methods=['POST'])
@login_required
@roles_required('admin')
def add_subject():
    subject_data = request.json
    
    try:
        with mysql_connection() as (mysql, cur):
            # Check if department exists
            cur.execute("SELECT department_id FROM departments WHERE department_id = %s", (subject_data.get('department_id'),))
            department = cur.fetchone()
            if department is None:
                return jsonify({'error': 'Department does not exist'}), 404
        
        
            # Insert subject
            cur.execute("""
                INSERT INTO subjects (user_id, name, department_id, type, weekly_hours) 
                VALUES (%s, %s, %s, %s, %s)
            """, 
                (session.get('user_id'),
                subject_data.get('subject_name'),
                subject_data.get("department_id"),
                subject_data.get('type_subject'),
                subject_data.get('weekly_hours'),
            ))
            cur.execute("""
                        INSERT INTO teacher_subjects (teacher_id, subject_id)
                        VALUES (%s, %s)
                        """,(
                            subject_data.get("teacher_id"),
                            cur.lastrowid,
                        ))
        
            # Log activity
            cur.execute("""
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id) 
                VALUES (%s, %s, %s, %s, %s)
            """, (
                session.get('user_id'),
                'CREATE',
                f'Added subject {subject_data.get("subject_name")}',
                'subject',
                cur.lastrowid,
            ))
            mysql.commit()
        
            return jsonify({'message': 'Subject added successfully', 'subject_id': cur.lastrowid}), 201
    
    except Exception as e:
        mysql.rollback()
        print(f"Error adding subject: {e}")
        return jsonify({'error': 'Failed to add subject: ' + str(e)}), 500
@app.route('/api/subjects/<int:subject_id>', methods=['GET'])
@login_required
def get_subject(subject_id):
    try:
        connection = get_mysql_connection()
        cur = connection.cursor(dictionary=True)  # Use dictionary cursor for named columns
        
        # Join with departments and teachers tables to get related data
        cur.execute("""
            SELECT s.subject_id, s.name, s.department_id, s.type, s.weekly_hours, 
                   d.name AS department_name,
                   ts.teacher_id,
                   t.teacher_name
            FROM subjects s
            LEFT JOIN departments d ON s.department_id = d.department_id
            LEFT JOIN teacher_subjects ts ON s.subject_id = ts.subject_id
            LEFT JOIN teachers t ON ts.teacher_id = t.teacher_id
            WHERE s.subject_id = %s and s.user_id=%s
        """, (subject_id, session.get("user_id"),))
        
        subject = cur.fetchone()
        
        if not subject:
            return jsonify({'error': 'Subject not found'}), 404
            
        return jsonify({
            'subject_id': subject['subject_id'],
            'name': subject['name'],
            'department_id': subject['department_id'],
            'department_name': subject['department_name'],
            'type': subject['type'],
            'weekly_hours': subject['weekly_hours'],
            'teacher_id': subject['teacher_id'],
            'teacher_name': subject['teacher_name']
        }), 200
    
    except Exception as e:
        return jsonify({'error': f'Failed to get subject: {str(e)}'}), 500
    
    finally:
        if 'cur' in locals():
            cur.close()
            mysql.close()
        if 'connection' in locals():
            connection.close()

@app.route('/api/subjects/<int:subject_id>', methods=['DELETE'])
@login_required
@roles_required('admin')
def delete_subject(subject_id):
    try:
        mysql = get_mysql_connection()
        cur = mysql.cursor(dictionary=True)
        cur.execute("SELECT * FROM subjects WHERE subject_id = %s", (subject_id,))
        subject = cur.fetchone()
        
        if not subject:
            return jsonify({'error': 'Subject not found'}), 404
        
        if subject['user_id'] != session.get("user_id"):
            return jsonify({'error': 'You do not have permission to delete this subject'}), 403
        
        # Start by deleting related records in teacher_subjects
        cur.execute("DELETE FROM teacher_subjects WHERE subject_id = %s", (subject_id,))
        
        # Then delete related records in class_subjects
        cur.execute("DELETE FROM class_subjects WHERE subject_id = %s", (subject_id,))
        
        # Finally delete the subject itself
        cur.execute("DELETE FROM subjects WHERE subject_id = %s", (subject_id,))
        
        # Log the activity
        cur.execute("""
                    INSERT INTO activity 
                    (user_id, action_type, description, entity_type, entity_id)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (
                        session.get('user_id'),
                        'DELETE',
                        f'Deleted subject {subject["name"]}',
                        'subject',
                        subject_id
                    ))
        
        mysql.commit()
        return jsonify({'message': 'Subject deleted successfully'}), 200
    except Exception as e:
        if mysql:
            mysql.rollback()
        return jsonify({'error': f'Failed to delete subject: {str(e)}'}), 500
    finally:
        if cur:
            cur.close()
        if mysql:
            mysql.close()
    
@app.route('/api/subjects/<int:subject_id>', methods=['PUT'])
@login_required
@roles_required('admin')
def update_subject(subject_id):
    data = request.json
    try:
        connection = get_mysql_connection()
        cur = connection.cursor(dictionary=True)  # Use dictionary cursor for named columns
        
        # Update the subject record
        cur.execute("""
                    UPDATE subjects
                    SET name = %s, department_id = %s, type = %s, weekly_hours = %s
                    WHERE subject_id = %s AND user_id = %s
                    """,
                    (
                        data['subject_name'],
                        data['department_id'],
                        data['type'],
                        data['weekly_hours'],
                        subject_id,
                        session.get("user_id"),
                    ))
        
        # Check if any rows were affected by the update
        if cur.rowcount == 0:
            return jsonify({'error': 'Subject not found or you do not have permission to update it'}), 404
        
        # Handle teacher association - first check if record exists
        cur.execute("""
                    SELECT * FROM teacher_subjects
                    WHERE subject_id = %s
                    """, (subject_id,))
        teacher_subject = cur.fetchone()  # This returns the row as a dictionary, not a list
        
        if teacher_subject:
            if teacher_subject['teacher_id'] != data['teacher_id']:
                # Update existing record
                cur.execute("""
                            UPDATE teacher_subjects
                            SET teacher_id = %s
                            WHERE subject_id = %s
                            """,
                            (
                                data['teacher_id'],
                                subject_id,
                            ))
        else:
            # Create new teacher_subject record if it doesn't exist
            cur.execute("""
                        INSERT INTO teacher_subjects (teacher_id, subject_id)
                        VALUES (%s, %s)
                        """,
                        (
                            data['teacher_id'],
                            subject_id,
                        ))
        cur.execute("""
            INSERT INTO activity (user_id, action_type, description, entity_type, entity_id) 
            VALUES (%s, %s, %s, %s, %s)
        """, (
            session.get('user_id'),
            'CREATE',
            f'Update subject {data['subject_name']}',
            'subject',
            subject_id,
        ))
        connection.commit()
        
        return jsonify({'message': 'Subject updated successfully'}), 200
    except Exception as e:
        # Rollback in case of error
        if connection:
            connection.rollback()
        return jsonify({'error': f'Failed to update subject: {str(e)}'}), 500
    finally:
        if cur:
            cur.close()
            mysql.close()
        if connection:
            connection.close()
    
@app.route('/api/classrooms', methods=['GET'])
@login_required
def get_classrooms():
    try:
        eff_user_id = get_effective_user_id()
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                SELECT classroom_id, room_number, CONCAT(UPPER(SUBSTRING(type, 1, 1)), LOWER(SUBSTRING(type, 2))) AS type, capacity, building, floor
                FROM classrooms
                WHERE user_id = %s
            """, (eff_user_id,))
            classrooms = cur.fetchall()
            return jsonify(classrooms)
    except Exception as e:
        print(f"Error fetching classrooms: {e}")
        return jsonify({'error': str(e)}), 500
@app.route('/api/classrooms/<int:classroom_id>', methods=['GET'])
@login_required
def get_classroom(classroom_id):
    try:
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                        SELECT classroom_id, room_number, type, capacity, building, floor
                        FROM classrooms
                        WHERE classroom_id = %s AND user_id = %s
                        """, (classroom_id, session.get('user_id')))
            classroom = cur.fetchone()
            if not classroom:
                return jsonify({'error': 'Classroom not found'}), 404
            return jsonify(classroom)
    except Exception as e:
        print(f"Error fetching classroom: {e}")
        return jsonify({'error': 'Failed to fetch classroom'}), 500
@app.route('/api/classrooms/<int:classroom_id>', methods=['DELETE'])
@login_required
@roles_required('admin')
def delete_classroom(classroom_id):
    
    try:
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                SELECT * FROM classrooms 
                WHERE classroom_id = %s AND user_id = %s
            """, (classroom_id, session.get('user_id')))
            classroom = cur.fetchone()
            if not classroom:
                return jsonify({'error': 'Classroom not found or not authorized'}), 404
            else:
                # Delete classroom
                cur.execute("""
                    DELETE FROM classrooms WHERE classroom_id = %s AND user_id = %s
                """, (classroom_id, session.get('user_id')))
                mysql.commit()
            
                # Log activity
                cur.execute("""
                    INSERT INTO activity (user_id, action_type, description, entity_type, entity_id) 
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    session.get('user_id'),
                    'DELETE',
                    f'Deleted classroom {classroom["room_number"]}',
                    'classroom',
                    classroom_id
                ))
                mysql.commit()
            return jsonify({'message': 'Classroom deleted successfully'}), 200
    except Exception as e:
        print(f"Error deleting classroom: {e}")
        mysql.rollback()
        return jsonify({'error': str(e)}), 500
@app.route('/api/classrooms/<int:classroom_id>', methods=['PUT'])
@login_required
@roles_required('admin')
def update_classroom(classroom_id):
    classroom_data = request.json
    try:
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                SELECT * FROM classrooms
                WHERE classroom_id = %s AND user_id = %s
            """, (classroom_id, session.get('user_id')))
            classroom = cur.fetchone()
            if not classroom:
                return jsonify({'error': 'Classroom not found or not authorized'}), 404
            else:
                # Convert floor to appropriate value
                floor_value = classroom_data.get('floor', '0')
            
                # Update classroom
                cur.execute("""
                    UPDATE classrooms
                    SET room_number = %s, capacity = %s, type = %s, building = %s, floor = %s
                    WHERE classroom_id = %s AND user_id = %s
                    """, (
                        classroom_data.get('room_number'), 
                        classroom_data.get('capacity'),
                        classroom_data.get('type'), 
                        classroom_data.get('building'), 
                        floor_value, 
                        classroom_id, 
                        session.get('user_id')
                    ))
                mysql.commit()
            
                # Log activity
                cur.execute("""
                    INSERT INTO activity (user_id, action_type, description, entity_type, entity_id)
                    VALUES (%s, %s, %s, %s, %s)
                    """, (
                        session.get('user_id'),
                        'UPDATE',
                        f'Updated classroom {classroom["room_number"]}',
                        'classroom',
                        classroom_id
                    ))
                mysql.commit()
                return jsonify({'message': 'Classroom updated successfully'}), 200
    except Exception as e:
        print(f"Error updating classroom: {e}")
        mysql.rollback()
        return jsonify({'error': str(e)}), 500
@app.route('/api/departments', methods=['GET'])
@login_required
def get_departments():
    try:
        eff_user_id = get_effective_user_id()
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                SELECT DISTINCT d.name, d.department_id
                FROM departments d
                LEFT JOIN teachers t
                ON t.department_id = d.department_id
                WHERE t.user_id = %s
            """, (eff_user_id,))
            departments = cur.fetchall()
            return jsonify(departments), 200
    except Exception as e:
        app.logger.error(f"Error fetching departments: {e}")
        return jsonify({'error': 'Failed to retrieve departments'}), 500
@app.route('/api/classrooms', methods=['POST'])
@login_required
@roles_required('admin')
def add_classroom():
    classroom_data = request.json
    
    try:
        with mysql_connection() as (mysql, cur):
            # Insert classroom
            cur.execute("""
                INSERT INTO classrooms (user_id, room_number, type, capacity, building, floor) 
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                session.get('user_id'),
                classroom_data.get('room_number'),
                classroom_data.get('type'),
                classroom_data.get('capacity'),
                classroom_data.get('building'),
                classroom_data.get('floor')
            ))
            mysql.commit()
        
            # Log activity
            cur.execute("""
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id) 
                VALUES (%s, %s, %s, %s, %s)
            """, (
                session.get('user_id'),
                'CREATE',
                f'Added classroom {classroom_data.get("room_number")}',
                'classroom',
                cur.lastrowid
            ))
            mysql.commit()
        
            return jsonify({'message': 'Classroom added successfully', 'classroom_id': cur.lastrowid}), 201
    
    except Exception as e:
        mysql.rollback()
        print(f"Error adding classroom: {e}")
        return jsonify({'error': 'Failed to add classroom: ' + str(e)}), 500
# Timetable Generation Routes
@app.route('/api/generate-timetable', methods=['POST'])
@login_required
@roles_required('admin')
def generate_timetable():
    timetable_params = request.json

    mysql = get_mysql_connection()
    cur = mysql.cursor(dictionary=True)
    
    # Extract form parameters
    timetable_name = timetable_params.get("timetable-name")
    timetable_description = timetable_params.get("timetable-description")
    start_date = timetable_params.get("start-date")
    end_date = timetable_params.get("end-date")
    start_time = timetable_params.get("school-start-time")
    end_time = timetable_params.get("school-end-time")
    lecture_duration = timetable_params.get("lecture-duration")
    break_duration = timetable_params.get("break-duration")
    classes_inc = timetable_params.get("classes")
    break_start = timetable_params.get("break-start-time")
    prioritize_teacher_preferences = timetable_params.get("prioritize-teacher-preferences")
    max_consecutive_lec = timetable_params.get("max-consecutive-classes")
    class_per_day = timetable_params.get("classes-per-day")
    allow_gaps = timetable_params.get("allow-gaps")
    
    # Get user_id from session
    user_id = session.get('user_id')
    
    # Corrected comprehensive query to fetch all required data
    cur.execute("""
        SELECT DISTINCT
            t.teacher_name,
            d.name as department_name,
            g.name as grade_name,
            s.name as subject_name,
            s.type as subject_type,
            s.weekly_hours,
            c.name as class_name,
            c.students_count,
            c.classroom_id,
            cr.room_number,
            cr.type as classroom_type,
            cr.capacity as classroom_capacity
        FROM teachers t
        JOIN teacher_subjects ts ON t.teacher_id = ts.teacher_id
        JOIN subjects s ON ts.subject_id = s.subject_id
        JOIN departments d ON s.department_id = d.department_id
        JOIN class_subjects cs ON s.subject_id = cs.subject_id
        JOIN classes c ON cs.class_id = c.class_id
        JOIN grades g ON c.grade_id = g.grade_id
        LEFT JOIN classrooms cr ON c.classroom_id = cr.classroom_id
        WHERE t.user_id = %s
        ORDER BY t.teacher_name, d.name, g.name, s.name
    """, (user_id,))
    
    results = cur.fetchall()
    
    # Process the results to create the required data structures
    from collections import defaultdict
    
    # Initialize data structures
    teachers = set()
    name_batch = set()
    name_of_subclass = set()
    classes_data = {}
    subject_assign_to_teacher_for_which_class = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    theory = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    practical = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    teacher_subjects_mapping = defaultdict(set)
    
    # Process results
    for row in results:
        teacher_name = row['teacher_name']
        dept_name = row['department_name']
        grade_name = row['grade_name']
        subject_name = row['subject_name']
        subject_type = row['subject_type']
        weekly_hours = row['weekly_hours'] or 0
        class_name = row['class_name']
        students_count = row['students_count']
        classroom_id = row['classroom_id']
        room_number = row['room_number']
        classroom_type = row['classroom_type']
        classroom_capacity = row['classroom_capacity']
        
        # Collect unique values
        teachers.add(teacher_name)
        name_batch.add(grade_name)
        name_of_subclass.add(dept_name)
        
        # Classes data (use class_name as key to avoid duplicates)
        if class_name not in classes_data:
            classes_data[class_name] = {
                'class_name': class_name,
                'grade_name': grade_name,
                'room_number': room_number,
                'students_count': students_count,
                'classroom_id': classroom_id,
                'classroom_type': classroom_type,
                'classroom_capacity': classroom_capacity
            }
        
        # Subject assignments (avoid duplicates)
        if subject_name not in subject_assign_to_teacher_for_which_class[teacher_name][dept_name][grade_name]:
            subject_assign_to_teacher_for_which_class[teacher_name][dept_name][grade_name].append(subject_name)
        
        # Teacher subjects mapping (use set to avoid duplicates)
        teacher_subjects_mapping[teacher_name].add(subject_name)
        
        # Theory and practical hours
        if subject_type == 'Theory':
            theory[teacher_name][dept_name][grade_name] += weekly_hours
        elif subject_type == 'Practical':
            practical[teacher_name][dept_name][grade_name] += weekly_hours
    
    # Convert to desired format
    teachers = sorted(list(teachers))
    name_batch = sorted(list(name_batch))
    name_of_subclass = sorted(list(name_of_subclass))
    no_batch = len(name_batch)
    
    # Convert defaultdict to regular dict and ensure all combinations exist
    final_subject_assign = {}
    final_theory = {}
    final_practical = {}
    
    for teacher in teachers:
        final_subject_assign[teacher] = {}
        final_theory[teacher] = {}
        final_practical[teacher] = {}
        
        for dept in name_of_subclass:
            final_subject_assign[teacher][dept] = {}
            final_theory[teacher][dept] = {}
            final_practical[teacher][dept] = {}
            
            for batch in name_batch:
                final_subject_assign[teacher][dept][batch] = subject_assign_to_teacher_for_which_class[teacher][dept][batch]
                final_theory[teacher][dept][batch] = theory[teacher][dept][batch]
                final_practical[teacher][dept][batch] = practical[teacher][dept][batch]
    
    # Format data for compatibility with existing code
    teacher_list = [{"teacher_name": name} for name in teachers]
    teacher_subjects_list = [{"teacher_name": teacher, "subject_list": list(subjects)} 
                           for teacher, subjects in teacher_subjects_mapping.items()]
    
    # Theory and practical count format
    theory_and_practical_list = []
    for teacher in teachers:
        for dept in name_of_subclass:
            for batch in name_batch:
                theory_hours = final_theory[teacher][dept][batch]
                practical_hours = final_practical[teacher][dept][batch]
                
                if theory_hours > 0:
                    theory_and_practical_list.append({
                        "teacher_name": teacher,
                        "subject_name": f"{dept} {batch}",
                        "subject_type": "Theory",
                        "theory_count": theory_hours,
                        "practical_count": 0
                    })
                
                if practical_hours > 0:
                    theory_and_practical_list.append({
                        "teacher_name": teacher,
                        "subject_name": f"{dept} {batch}",
                        "subject_type": "Practical",
                        "theory_count": 0,
                        "practical_count": practical_hours
                    })
    
    classes_list = list(classes_data.values())
    
    # Fetch Teacher Availability
    cur.execute("""
        SELECT t.teacher_name, ta.day_of_week, 
               TIME_FORMAT(ta.start_time, '%H:%i') as start_time, 
               TIME_FORMAT(ta.end_time, '%H:%i') as end_time, 
               ta.preference_level
        FROM teacher_availability ta
        JOIN teachers t ON ta.teacher_id = t.teacher_id
        WHERE t.user_id = %s
    """, (user_id,))
    availability_results = cur.fetchall()
    
    teacher_availability_dict = defaultdict(list)
    for avail in availability_results:
        teacher_availability_dict[avail['teacher_name']].append({
            'day': avail['day_of_week'],
            'start': avail['start_time'],
            'end': avail['end_time'],
            'status': avail['preference_level']
        })

    # Get additional classroom information
    cur.execute("""
        SELECT classroom_id, room_number, type, capacity, building, floor
        FROM classrooms
        WHERE user_id = %s
    """, (user_id,))
    classroom_additional = cur.fetchall()
    
    # Create the timetable parameters dictionary
    main_timetable_params_dict = {
        "teachers": teachers,
        "name_batch": name_batch,
        "name_of_subclass": name_of_subclass,
        "subject_assign_to_teacher_for_which_class": final_subject_assign,
        "classrooms_additional": classroom_additional,
        "no_batch": no_batch,
        "theory": final_theory,
        "practical": final_practical,
        "teacher": teacher_list,
        "subjects": teacher_subjects_list,
        "theory_and_practical": theory_and_practical_list,
        "classrooms": classes_list,
        "classes": classes_list,
        "teacher_availability": dict(teacher_availability_dict),
        "start_date": start_date,
        "end_date": end_date,
        "start_time": start_time,
        "end_time": end_time,
        "lecture_duration": lecture_duration,
        "break_duration": break_duration,
        "classes_inc": classes_inc,
        "break_start": break_start,
        "prioritize_teacher_preferences": prioritize_teacher_preferences,
        "max_consecutive_lec": max_consecutive_lec,
        "class_per_day": class_per_day,
        "allow_gaps": allow_gaps
    }
    
    try:
        # Generate timetable using the comprehensive data
        output = Generate_test_timetable(main_timetable_params_dict)
        
        # Convert the output to JSON format
        if isinstance(output, dict):
            output_json = json.dumps(output, indent=4, ensure_ascii=False)
        else:
            output_json = str(output)
        
        # Add timetable metadata to main_timetable_params_dict
        main_timetable_params_dict["timetable_description"] = timetable_description
        main_timetable_params_dict["timetable_name"] = timetable_name
        main_timetable_params_dict["timetable_data"] = output
        main_timetable_params_dict["generated_at"] = datetime.now().isoformat()
        
        print("Timetable generation completed successfully")
        
        # Save to JSON files
        json_main_timetable = json.dumps(main_timetable_params_dict, indent=4, ensure_ascii=False, default=str)
        
        # Update/Insert school configuration in MySQL
        try:
            # Check if school config exists for this user
            cur.execute("SELECT config_id FROM school_config WHERE user_id = %s", (user_id,))
            existing_config = cur.fetchone()
            
            # Prepare departments JSON
            departments_json = json.dumps(name_of_subclass) if name_of_subclass else json.dumps([])
            
            if existing_config:
                # Update existing configuration
                cur.execute("""
                    UPDATE school_config 
                    SET start_time = %s, end_time = %s, department_include = %s, 
                        lecture_duration = %s, break_duration = %s, lunch_break_start = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE user_id = %s
                """, (start_time, end_time, departments_json, lecture_duration, 
                      break_duration, break_start, user_id))
                
                config_action = "Updated"
            else:
                # Insert new configuration
                cur.execute("""
                    INSERT INTO school_config 
                    (user_id, school_name, start_time, end_time, department_include, 
                     lecture_duration, break_duration, lunch_break_start)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (user_id, "School", start_time, end_time, departments_json, 
                      lecture_duration, break_duration, break_start))
                
                config_action = "Created"
            
            # Log activity in MySQL
            cur.execute("""
                INSERT INTO activity 
                (user_id, action_type, description, entity_type, timestamp)
                VALUES (%s, %s, %s, %s, %s)
            """, (user_id, "TIMETABLE_GENERATED", 
                  f"Generated timetable: {timetable_name}", "TIMETABLE", 
                  datetime.now()))
            
            # Log school config activity
            cur.execute("""
                INSERT INTO activity 
                (user_id, action_type, description, entity_type, timestamp)
                VALUES (%s, %s, %s, %s, %s)
            """, (user_id, "SCHOOL_CONFIG_UPDATED", 
                  f"{config_action} school configuration", "SCHOOL_CONFIG", 
                  datetime.now()))
            
            mysql.commit()
            
        except Exception as sql_error:
            print(f"Error updating SQL database: {str(sql_error)}")
            mysql.rollback()
        
        # Save to MongoDB
        try:
            mongo_db = get_mongodb_connection()
            timetables_collection = mongo_db["timetables"]
            
            # Prepare MongoDB document
            mongodb_document = {
                "user_id": user_id,
                "timetable_name": timetable_name,
                "timetable_description": timetable_description,
                "generated_at": datetime.now(),
                "parameters": {
                    "start_date": start_date,
                    "end_date": end_date,
                    "start_time": start_time,
                    "end_time": end_time,
                    "lecture_duration": lecture_duration,
                    "break_duration": break_duration,
                    "classes_inc": classes_inc,
                    "break_start": break_start,
                    "prioritize_teacher_preferences": prioritize_teacher_preferences,
                    "max_consecutive_lec": max_consecutive_lec,
                    "class_per_day": class_per_day,
                    "allow_gaps": allow_gaps
                },
                "data": {
                    "teachers": teachers,
                    "name_batch": name_batch,
                    "name_of_subclass": name_of_subclass,
                    "classes": classes_list,
                    "theory_and_practical": theory_and_practical_list
                },
                "generated_timetable": output,
                "status": "active"
            }
            
            # Insert into MongoDB
            result = timetables_collection.insert_one(mongodb_document)
            mongodb_id = str(result.inserted_id)
            
            print(f"Timetable saved to MongoDB with ID: {mongodb_id}")
            
        except Exception as mongo_error:
            print(f"Error saving to MongoDB: {str(mongo_error)}")
            mongodb_id = None
        
        # Close database connections
        cur.close()
        mysql.close()
        
        return jsonify({
            "success": True,
            "message": "Timetable generated successfully",
            "timetable": output_json,
            "mongodb_id": mongodb_id,
            "files_created": [
                "Output.json",
                "Timetable_Data_final_output.json"
            ],
            "database_updates": {
                "mysql": "school_config and activity tables updated",
                "mongodb": "timetable document saved"
            }
        })
        
    except Exception as e:
        # Close database connections in case of error
        cur.close()
        mysql.close()
        
        print(f"Error generating timetable: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Error generating timetable: {str(e)}",
            "data": None,
            "timetable": None
        }), 500

from bson import ObjectId
def convert_bson(doc):
    """ Convert MongoDB BSON types to JSON serializable """
    for key, value in doc.items():
        if isinstance(value, ObjectId):
            doc[key] = str(value)
        elif isinstance(value, datetime):
            doc[key] = value.isoformat()
        elif isinstance(value, dict):
            doc[key] = convert_bson(value)
        elif isinstance(value, list):
            doc[key] = [convert_bson(item) if isinstance(item, dict) else item for item in value]
    return doc

   
@app.route('/api/view-timetables', methods=['GET'])
@login_required
def get_timetables():
    eff_user_id = get_effective_user_id()
    mongo_db = get_mongodb_connection()
    timetables_cursor = mongo_db["timetables"].find({"user_id": eff_user_id})
    
    timetables = [convert_bson(doc) for doc in timetables_cursor]
    generated_timetables = [json.loads(t['generated_timetable']) for t in timetables if 'generated_timetable' in t]

    return jsonify({"Time-Params":timetables, "Generated-Value":generated_timetables})

# @app.route('/api/timetables/<timetable_id>', methods=['GET'])
# @login_required
# def get_timetable(timetable_id):
#     from bson.objectid import ObjectId
    
#     mongo_db = get_mongodb_connection()
#     timetable = mongo_db["timetables"].find_one({"_id": ObjectId(timetable_id)})
    
#     if not timetable:
#         return jsonify({"error": "Timetable not found"}), 404
    
#     # Convert ObjectId to string for JSON serialization
#     timetable["_id"] = str(timetable["_id"])
#     timetable_value=timetable['generated_timetable']
#     # value dumps as json
#     timetable_value=json.dumps(timetable_value, indent=4, ensure_ascii=False, default=str)
#     return jsonify(timetable, timetable_value)

# Absence Management Routes
@app.route('/api/absences', methods=['GET'])
@login_required
def get_absences():
    try:
        eff_user_id = get_effective_user_id()
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                SELECT ta.absence_id, ta.start_date, ta.end_date, ta.reason, ta.status,
                       t.teacher_name, t.teacher_id
                FROM teacher_absences ta
                JOIN teachers t ON ta.teacher_id = t.teacher_id
                WHERE t.user_id = %s
                ORDER BY ta.start_date DESC
            """, (eff_user_id,))
        
            absences = cur.fetchall()
        
            # Convert date objects to strings and add affected count
            for absence in absences:
                sd = absence['start_date'].strftime('%Y-%m-%d')
                ed = absence['end_date'].strftime('%Y-%m-%d')
                absence['start_date'] = sd
                absence['end_date'] = ed
                
                # Fetch affected classes count
                affected = get_affected_classes(absence['teacher_name'], sd, ed, eff_user_id)
                absence['affected_count'] = len(affected)
        
            return jsonify(absences)
    
    except Exception as e:
        app.logger.error(f"Error fetching absences: {e}")
        return jsonify({'error': str(e)}), 500
@app.route('/api/absences', methods=['POST'])
@login_required
@roles_required('admin')
def add_absence():
    absence_data = request.json
    
    try:
        with mysql_connection() as (mysql, cur):
            # Verify teacher belongs to user
            cur.execute("""
                SELECT teacher_id FROM teachers 
                WHERE teacher_id = %s AND user_id = %s
            """, (absence_data.get('teacher_id'), session.get('user_id')))
        
            if not cur.fetchone():
                return jsonify({'error': 'Teacher not found or not authorized'}), 404
        
            # Add absence
            cur.execute("""
                INSERT INTO teacher_absences (teacher_id, start_date, end_date, reason, status) 
                VALUES (%s, %s, %s, %s, %s)
            """, (
                absence_data.get('teacher_id'),
                absence_data.get('start_date'),
                absence_data.get('end_date'),
                absence_data.get('reason'),
                absence_data.get('status', 'pending')
            ))
            mysql.commit()
        
            # Log activity
            cur.execute("""
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id) 
                VALUES (%s, %s, %s, %s, %s)
            """, (
                session.get('user_id'),
                'CREATE',
                f'Added teacher absence',
                'absence',
                cur.lastrowid
            ))
            mysql.commit()
        
            return jsonify({'message': 'Absence recorded successfully', 'absence_id': cur.lastrowid}), 201
    
    except Exception as e:
        mysql.rollback()
        print(f"Error adding absence: {e}")
        return jsonify({'error': 'Failed to add absence: ' + str(e)}), 500
@app.route('/api/absences/<int:absence_id>', methods=['GET'])
@login_required
def get_absence(absence_id: int):
    try:
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                SELECT ta.*, t.teacher_name 
                FROM teacher_absences ta
                JOIN teachers t ON ta.teacher_id = t.teacher_id
                WHERE ta.absence_id = %s AND t.user_id = %s
            """, (absence_id, session.get('user_id')))
            absence = cur.fetchone()
            if not absence:
                return jsonify({'error': 'Absence not found or not authorized'}), 404
            
            # Convert date objects to strings
            absence['start_date'] = absence['start_date'].strftime('%Y-%m-%d')
            absence['end_date'] = absence['end_date'].strftime('%Y-%m-%d')
            if 'created_at' in absence and absence['created_at']:
                absence['created_at'] = absence['created_at'].isoformat()
            if 'updated_at' in absence and absence['updated_at']:
                absence['updated_at'] = absence['updated_at'].isoformat()
                
            return jsonify(absence)
    except Exception as e:
        app.logger.error(f"Error fetching absence {absence_id}: {e}")
        return jsonify({'error': 'Failed to fetch absence'}), 500

@app.route('/api/absences/<int:absence_id>', methods=['DELETE'])
@login_required
@roles_required('admin')
def delete_absence(absence_id: int):
    try:
        with mysql_connection() as (mysql, cur):
            # Ensure the absence belongs to a teacher owned by this user
            cur.execute("""
                SELECT a.absence_id FROM teacher_absences a
                JOIN teachers t ON a.teacher_id = t.teacher_id
                WHERE a.absence_id = %s AND t.user_id = %s
            """, (absence_id, session.get('user_id')))
            if not cur.fetchone():
                return jsonify({'error': 'Absence not found or not authorized'}), 404

            cur.execute("DELETE FROM teacher_absences WHERE absence_id = %s", (absence_id,))
            mysql.commit()

            # Activity log
            cur.execute("""
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id)
                VALUES (%s, 'DELETE', 'Deleted teacher absence', 'absence', %s)
            """, (session.get('user_id'), absence_id))
            mysql.commit()

            return jsonify({'message': 'Absence deleted successfully'})
    except Exception as e:
        mysql.rollback()
        app.logger.error(f"Error deleting absence {absence_id}: {e}")
        return jsonify({'error': 'Failed to delete absence'}), 500

@app.route('/api/absences/<int:absence_id>', methods=['PUT'])
@login_required
def update_absence(absence_id: int):
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Request body required'}), 400

    try:
        with mysql_connection() as (mysql, cur):
            # Ownership check
            cur.execute("""
                SELECT a.absence_id FROM teacher_absences a
                JOIN teachers t ON a.teacher_id = t.teacher_id
                WHERE a.absence_id = %s AND t.user_id = %s
            """, (absence_id, session.get('user_id')))
            if not cur.fetchone():
                return jsonify({'error': 'Absence not found or not authorized'}), 404

            # Build update dynamically for all provided fields
            fields, values = [], []
            
            if 'teacher_id' in data:
                fields.append("teacher_id = %s")
                values.append(data['teacher_id'])
            
            if 'start_date' in data:
                fields.append("start_date = %s")
                values.append(data['start_date'])
                
            if 'end_date' in data:
                fields.append("end_date = %s")
                values.append(data['end_date'])
                
            if 'reason' in data:
                fields.append("reason = %s")
                values.append(data['reason'])
                
            if 'status' in data:
                allowed_statuses = {'pending', 'resolved'}
                if data['status'] not in allowed_statuses:
                    return jsonify({'error': f"Invalid status. Allowed: {allowed_statuses}"}), 422
                fields.append("status = %s")
                values.append(data['status'])

            if not fields:
                return jsonify({'error': 'No updatable fields provided'}), 400

            values.append(absence_id)
            cur.execute(
                f"UPDATE teacher_absences SET {', '.join(fields)} WHERE absence_id = %s",
                values
            )
            mysql.commit()

            cur.execute("""
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id)
                VALUES (%s, 'UPDATE', 'Updated teacher absence', 'absence', %s)
            """, (session.get('user_id'), absence_id))
            mysql.commit()

            return jsonify({'message': 'Absence updated successfully'})
    except Exception as e:
        mysql.rollback()
        app.logger.error(f"Error updating absence {absence_id}: {e}")
        return jsonify({'error': 'Failed to update absence'}), 500
# ── Absence Resolution Endpoints ─────────────────────────────────────────────

def get_affected_classes(teacher_name, start_date, end_date, user_id):
    """Finds all classes/lectures for a teacher between start_date and end_date."""
    mongo_db = get_mongodb_connection()
    # Get all timetables for this user
    timetables = mongo_db["timetables"].find({"user_id": user_id})
    
    affected = []
    days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    
    # Iterate through all generated timetables to find the teacher's schedule
    for doc in timetables:
        generated = doc.get('generated_timetable', {})
        if isinstance(generated, str):
            try:
                generated = json.loads(generated)
            except:
                continue
        
        for class_name, schedule in generated.items():
            for day, slots in schedule.items():
                for slot, details in slots.items():
                    if teacher_name.lower() in details.get('subject', '').lower():
                        affected.append({
                            'class_name': class_name,
                            'day': day,
                            'slot': slot,
                            'subject': details.get('subject'),
                            'classroom': details.get('classroom')
                        })
    
    # Deduplicate by subject, class, day, slot
    unique_affected = []
    seen = set()
    for item in affected:
        key = (item['class_name'], item['day'], item['slot'])
        if key not in seen:
            unique_affected.append(item)
            seen.add(key)
            
    return unique_affected

@app.route('/api/absences/<int:absence_id>/affected-classes', methods=['GET'])
@login_required
def get_absence_affected_classes(absence_id: int):
    try:
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                SELECT ta.*, t.teacher_name 
                FROM teacher_absences ta
                JOIN teachers t ON ta.teacher_id = t.teacher_id
                WHERE ta.absence_id = %s AND t.user_id = %s
            """, (absence_id, session.get('user_id')))
            absence = cur.fetchone()
            if not absence:
                return jsonify({'error': 'Absence not found'}), 404
            
            classes = get_affected_classes(
                absence['teacher_name'], 
                absence['start_date'], 
                absence['end_date'], 
                session.get('user_id')
            )
            return jsonify({
                'teacher_name': absence['teacher_name'],
                'affected_classes': classes
            })
    except Exception as e:
        app.logger.error(f"Error fetching affected classes: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/absences/<int:absence_id>/auto-suggest', methods=['GET'])
@login_required
def auto_suggest_replacement(absence_id: int):
    """Return ranked list of available teachers who can cover the absent teacher's classes."""
    user_id = session.get('user_id')
    try:
        with mysql_connection() as (mysql, cur):
            # Fetch the absence record and verify ownership
            cur.execute(
                """
                SELECT ta.absence_id, ta.teacher_id, ta.start_date, ta.end_date,
                       t.teacher_name, t.department_name
                FROM teacher_absences ta
                JOIN teachers t ON ta.teacher_id = t.teacher_id
                WHERE ta.absence_id = %s AND t.user_id = %s
                """,
                (absence_id, user_id),
            )
            absence = cur.fetchone()
            if absence is None:
                return jsonify({'error': 'Absence not found or not authorized'}), 404

            # Find other teachers belonging to this user who are not absent during the period
            cur.execute(
                """
                SELECT t.teacher_id, t.teacher_name, t.department_name, t.weekly_hours
                FROM teachers t
                WHERE t.user_id = %s
                  AND t.teacher_id != %s
                  AND t.teacher_id NOT IN (
                      SELECT ta2.teacher_id FROM teacher_absences ta2
                      WHERE ta2.start_date <= %s AND ta2.end_date >= %s
                        AND ta2.status != 'resolved'
                  )
                ORDER BY
                    -- Same department gets priority score of 1 (lower = better)
                    CASE WHEN t.department_name = %s THEN 0 ELSE 1 END,
                    t.weekly_hours ASC
                LIMIT 10
                """,
                (
                    user_id,
                    absence['teacher_id'],
                    absence['end_date'],
                    absence['start_date'],
                    absence['department_name'],
                ),
            )
            suggestions = cur.fetchall()

            return jsonify({
                'absence_id': absence_id,
                'absent_teacher': absence['teacher_name'],
                'period': {
                    'start': str(absence['start_date']),
                    'end': str(absence['end_date']),
                },
                'suggestions': suggestions,
            })

    except Exception as e:
        app.logger.error(f"Error fetching auto-suggestions for absence {absence_id}: {e}")
        return jsonify({'error': 'Failed to fetch suggestions'}), 500
@app.route('/api/absences/<int:absence_id>/manual-assign', methods=['POST'])
@login_required
@roles_required('admin')
def manual_assign_replacement(absence_id: int):
    """Assign a specific replacement teacher to an absence and mark it resolved."""
    user_id = session.get('user_id')
    data = request.get_json()
    if not data or 'replacement_teacher_id' not in data:
        return jsonify({'error': 'replacement_teacher_id is required'}), 400

    replacement_id = data['replacement_teacher_id']
    try:
        with mysql_connection() as (mysql, cur):
            # Verify absence ownership
            cur.execute(
                """
                SELECT ta.absence_id FROM teacher_absences ta
                JOIN teachers t ON ta.teacher_id = t.teacher_id
                WHERE ta.absence_id = %s AND t.user_id = %s
                """,
                (absence_id, user_id),
            )
            if cur.fetchone() is None:
                return jsonify({'error': 'Absence not found or not authorized'}), 404

            # Verify replacement teacher belongs to same user
            cur.execute(
                "SELECT teacher_name FROM teachers WHERE teacher_id = %s AND user_id = %s",
                (replacement_id, user_id),
            )
            replacement = cur.fetchone()
            if replacement is None:
                return jsonify({'error': 'Replacement teacher not found'}), 404

            # Store resolution: mark resolved and embed replacement name in reason
            note = f"[Replaced by: {replacement['teacher_name']}]"
            cur.execute(
                """
                UPDATE teacher_absences
                SET status = 'resolved',
                    reason = CASE
                        WHEN reason IS NULL OR reason = '' THEN %s
                        ELSE CONCAT(reason, ' ', %s)
                    END
                WHERE absence_id = %s
                """,
                (note, note, absence_id),
            )
            mysql.commit()

            # Activity log
            cur.execute(
                """
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id)
                VALUES (%s, 'UPDATE', %s, 'absence', %s)
                """,
                (
                    user_id,
                    f"Manually assigned {replacement['teacher_name']} as replacement",
                    absence_id,
                ),
            )
            mysql.commit()

            return jsonify({
                'message': f"Assigned {replacement['teacher_name']} as replacement",
                'absence_id': absence_id,
                'status': 'resolved',
            })

    except Exception as e:
        mysql.rollback()
        app.logger.error(f"Error assigning replacement for absence {absence_id}: {e}")
        return jsonify({'error': 'Failed to assign replacement'}), 500
@app.route('/api/absences/<int:absence_id>/reschedule', methods=['POST'])
@login_required
@roles_required('admin')
def reschedule_absence(absence_id: int):
    """Record a reschedule resolution: new date range, then mark absence resolved."""
    user_id = session.get('user_id')
    data = request.get_json()
    if not data or not data.get('new_start_date') or not data.get('new_end_date'):
        return jsonify({'error': 'new_start_date and new_end_date are required'}), 400

    new_start = data['new_start_date']
    new_end = data['new_end_date']
    if new_end < new_start:
        return jsonify({'error': 'new_end_date must be on or after new_start_date'}), 422

    try:
        with mysql_connection() as (mysql, cur):
            # Verify absence ownership
            cur.execute(
                """
                SELECT ta.absence_id FROM teacher_absences ta
                JOIN teachers t ON ta.teacher_id = t.teacher_id
                WHERE ta.absence_id = %s AND t.user_id = %s
                """,
                (absence_id, user_id),
            )
            if cur.fetchone() is None:
                return jsonify({'error': 'Absence not found or not authorized'}), 404

            # Mark absence resolved and append reschedule note to reason field
            reschedule_note = data.get('note', f'Rescheduled to {new_start} – {new_end}')
            note_tag = f"[{reschedule_note}]"
            cur.execute(
                """
                UPDATE teacher_absences
                SET status = 'resolved',
                    reason = CASE
                        WHEN reason IS NULL OR reason = '' THEN %s
                        ELSE CONCAT(reason, ' ', %s)
                    END
                WHERE absence_id = %s
                """,
                (note_tag, note_tag, absence_id),
            )
            mysql.commit()

            # Activity log
            cur.execute(
                """
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id)
                VALUES (%s, 'UPDATE', %s, 'absence', %s)
                """,
                (user_id, reschedule_note, absence_id),
            )
            mysql.commit()

            return jsonify({
                'message': 'Absence rescheduled successfully',
                'absence_id': absence_id,
                'rescheduled_to': {'start': new_start, 'end': new_end},
                'status': 'resolved',
            })

    except Exception as e:
        mysql.rollback()
        app.logger.error(f"Error rescheduling absence {absence_id}: {e}")
        return jsonify({'error': 'Failed to reschedule absence'}), 500
@app.route('/api/grades', methods=['GET'])
@login_required
def get_grades():
    try:
        eff_user_id = get_effective_user_id()
        with mysql_connection() as (mysql, cur):
            cur.execute("""
                        SELECT DISTINCT g.grade_id, g.name
                        FROM grades g
                        JOIN classes c ON g.grade_id = c.grade_id
                        WHERE c.user_id = %s
                        """, (eff_user_id,))
            
            grades=cur.fetchall()
            
            return jsonify(grades)
    except Exception as e:
        print(f"Error fetching grades: {e}")
        return jsonify({'error': 'Failed to fetch grades: ' + str(e)}), 500
        

@app.route('/api/profile', methods=['GET', 'POST'])
@login_required
def profile():
    user_id = session.get('user_id')
    
    if request.method == 'POST':
        user_data = request.get_json()
        
        try:
            with mysql_connection() as (mysql, cursor):
                # Validate required fields
                if not all(key in user_data for key in ['name', 'email']):
                    return jsonify({'error': 'Name and email are required'}), 400
            
                # Check if email is taken by another user
                cursor.execute("""
                    SELECT user_id FROM users 
                    WHERE email = %s AND user_id != %s
                """, (user_data['email'], user_id))
            
                if cursor.fetchone():
                    return jsonify({'error': 'Email is already taken'}), 409
            
                # Update user profile
                update_query = """
                    UPDATE users 
                    SET name = %s, email = %s, updated_at = %s 
                    WHERE user_id = %s
                """
                cursor.execute(update_query, (
                    user_data['name'],
                    user_data['email'],
                    datetime.now(),
                    user_id
                ))
            
                mysql.commit()
            
                # Update session
                session['user_name'] = user_data['name']
            
                return jsonify({
                    'message': 'Profile updated successfully',
                    'user': {
                        'name': user_data['name'],
                        'email': user_data['email']
                    }
                })
            
        except Exception as e:
            mysql.rollback()
            print(f"Error updating profile: {e}")
            return jsonify({'error': 'Failed to update profile'}), 500
    elif request.method == "GET":
        
        try:
            with mysql_connection() as (mysql, cursor):
                query = "SELECT user_id, name, email, role, created_at FROM users WHERE user_id = %s"
                cursor.execute(query, (user_id,))
                user_data = cursor.fetchone()
            
                if not user_data:
                    return jsonify({'error': 'User not found'}), 404
                # user_profile_picture = None
                # if user_data['profile_picture']!=None:
                #     user_profile_picture = user_data['profile_picture'].decode('utf-8')
            
                return jsonify({
                    'stats': {
                        'user_id': user_data['user_id'],
                        'name': user_data['name'],
                        'email': user_data['email'],
                        'role': user_data['role'],
                        # 'profile_picture':user_profile_picture,
                        'joined': user_data['created_at'].strftime('%Y-%m-%d')
                    }
                })
            
        except Exception as e:
            print(f"Error fetching profile: {e}")
            return jsonify({'error': 'Failed to fetch profile'}), 500
@app.route('/api/change-password', methods=['POST'])
@login_required
def update_password():

    try:
        with mysql_connection() as (mysql, cursor):
            user_id = session.get('user_id')
            if user_id is None:
                return jsonify({'error': 'User not authenticated'}), 401

            # Get passwords from request
            new_password = request.json.get('new_password')
            current_password = request.json.get('current_password')
            if not new_password:
                return jsonify({'error': 'New password is required'}), 400
            if not current_password:
                return jsonify({'error': 'Current password is required'}), 400

            cursor.execute("""
                SELECT password_hash FROM users
                WHERE user_id = %s
            """, (user_id,))
            user = cursor.fetchone()
            if user is None:
                return jsonify({'error': 'User not found'}), 404

            # Verify current password using check_password_hash
            if not check_password_hash(user['password_hash'], current_password):
                return jsonify({'error': 'Current password is incorrect'}), 401

            # Hash the new password (default pbkdf2:sha256)
            hashed_password = generate_password_hash(new_password)

            # Update password in database
            cursor.execute("""
                UPDATE users
                SET password_hash = %s, updated_at = %s
                WHERE user_id = %s
            """, (hashed_password, datetime.now(), user_id))

            # Log activity
            cursor.execute("""
                INSERT INTO activity (user_id, action_type, description, entity_type, entity_id)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                user_id,
                'UPDATE',
                'Updated password',
                'user',
                user_id
            ))
            mysql.commit()
            return jsonify({
                'message': 'Password updated successfully',
                'timestamp': datetime.now().isoformat()
            })

    except Exception as e:
        mysql.rollback()
        app.logger.error(f"Error updating password: {e}")
        return jsonify({'error': 'Failed to update password'}), 500

@app.route('/api/profile/profile_picture', methods=['POST'])
@login_required
def update_profile_picture():
    try:
        with mysql_connection() as (mysql, cursor):
            user_id = session.get('user_id')
            if not user_id:
                return jsonify({'error': 'User not authenticated'}), 401
        
            # Check if file was uploaded
            if 'profile_picture' not in request.files:
                return jsonify({'error': 'No file provided'}), 400
            
            file = request.files['profile_picture']
        
            # Check if filename is empty
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400
            
            # Check file type
            if not file.content_type.startswith('image/'):
                return jsonify({'error': 'File must be an image'}), 400
        
            # Check file size (5MB limit)
            file_content = file.read()
            file_size = len(file_content)
            file.seek(0)  # Reset file pointer to beginning
        
            MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB in bytes
            if file_size > MAX_FILE_SIZE:
                return jsonify({'error': 'File size must be less than 5MB'}), 400
            
            # Additional validation for image content
            try:
                img = Image.open(file)
                img.verify()  # Verify it's a valid image
                file.seek(0)  # Reset file position after verification
            except Exception:
                return jsonify({'error': 'Invalid image file'}), 400
            
            # Read file content as binary data for BLOB storage
            image_data = file.read()
        
            # Connect to database
            mysql = get_mysql_connection()
            cursor = mysql.cursor()
        
            # Update profile picture in database (BLOB data type)
            query = "UPDATE users SET profile_picture = %s WHERE user_id = %s"
            cursor.execute(query, (image_data, user_id))
            mysql.commit()
        
            return jsonify({
                'message': 'Profile picture updated successfully',
                'timestamp': datetime.now().isoformat()
            })
    except Exception as e:
        if mysql:
            mysql.rollback()
        app.logger.error(f"Error updating profile picture: {str(e)}")
        return jsonify({'error': 'Failed to update profile picture'}), 500
# @app.route('/profile<setting', methods=['GET', 'POST'])
# @login_required()
# def toggle_theme():
#     mysql = get_mysql_connection()
#     cur = mysql.cursor()
#     user_id=session.get('user_id')
#     cur.execute('SELECT * FROM user_preferences WHERE user_id=%s')
#     user_preferences = cur.fetchone()
#     if request.method == "POST":
#         setting = request.form.get('setting')
#         if setting == 'theme':
#             if user_preferences['theme'] == 'light':
#                 cur.execute('UPDATE user_preferences SET theme = %s WHERE user_id = %s', ('TRUE', user_id))
#             else:
#                 cur.execute('UPDATE user_preferences SET theme = %s WHERE user_id = %s', (''))
#                 mysql.commit()
#                 return jsonify({'success': 'Theme toggled'}), 200
#             return jsonify({'error': 'Invalid setting'}), 400
    
#     elif request.method=="GET":
#         if user_preferences['dark_mode'] == '':
#             return jsonify({'dark_mode': 'dark'}), 200
        
    
    # if session.get('theme') == 'dark':
    #     session['theme'] = 'light'
    # else:
    #     session['theme'] = 'dark'
    # return redirect(request.referrer or url_for('home'))


@app.route('/api/export/excel', methods=['GET'])
@login_required
def export_excel():
    export_type = request.args.get('type')
    user_id = session.get('user_id')
    filter_value = request.args.get('filter_value')
    
    output = io.BytesIO()
    filename = f"{export_type}_export.xlsx"
    
    try:
        if export_type == 'teachers':
            with mysql_connection() as (mysql, cur):
                query = """
                    SELECT t.teacher_name, t.email, t.phone, t.weekly_hours, d.name AS department
                    FROM teachers t
                    JOIN departments d ON t.department_id = d.department_id
                    WHERE t.user_id = %s
                """
                params = [user_id]
                if filter_value and filter_value != 'All Department':
                    query += " AND d.name = %s"
                    params.append(filter_value)
                
                cur.execute(query, tuple(params))
                data = cur.fetchall()
                df = pd.DataFrame(data)
                if not df.empty:
                    df.columns = ['Name', 'Email', 'Phone', 'Weekly Hours', 'Department']
                
        elif export_type == 'classes':
            with mysql_connection() as (mysql, cur):
                query = """
                    SELECT c.name, g.name AS grade, c.students_count, cr.room_number
                    FROM classes c
                    LEFT JOIN grades g ON c.grade_id = g.grade_id
                    LEFT JOIN classrooms cr ON c.classroom_id = cr.classroom_id
                    WHERE c.user_id = %s
                """
                params = [user_id]
                if filter_value and filter_value != 'All':
                    query += " AND g.name = %s"
                    params.append(filter_value)
                
                cur.execute(query, tuple(params))
                data = cur.fetchall()
                df = pd.DataFrame(data)
                if not df.empty:
                    df.columns = ['Class Name', 'Grade', 'Students Count', 'Classroom']

        elif export_type == 'subjects':
            with mysql_connection() as (mysql, cur):
                query = """
                    SELECT s.name, s.type, s.weekly_hours, d.name AS department
                    FROM subjects s
                    JOIN departments d ON s.department_id = d.department_id
                    WHERE s.user_id = %s
                """
                params = [user_id]
                if filter_value:
                    query += " AND d.name = %s"
                    params.append(filter_value)
                
                cur.execute(query, tuple(params))
                data = cur.fetchall()
                df = pd.DataFrame(data)
                if not df.empty:
                    df.columns = ['Subject Name', 'Type', 'Weekly Hours', 'Department']

        elif export_type == 'timetable':
            timetable_id = request.args.get('id')
            category_type = request.args.get('category_type', 'class')
            
            mongo_db = get_mongodb_connection()
            # If no ID, get the latest one
            if not timetable_id or timetable_id == 'undefined':
                timetable_doc = mongo_db["timetables"].find_one({"user_id": user_id}, sort=[("generated_at", -1)])
            else:
                timetable_doc = mongo_db["timetables"].find_one({"_id": ObjectId(timetable_id), "user_id": user_id})
            
            if not timetable_doc:
                return jsonify({'error': 'Timetable not found'}), 404
            
            generated_timetable = timetable_doc['generated_timetable']
            if isinstance(generated_timetable, str):
                generated_timetable = json.loads(generated_timetable)
            
            lecture_duration = int(timetable_doc.get('parameters', {}).get('lecture_duration', 45))
            
            # Prepare Excel writer with multiple sheets if needed
            sheet_added = False
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                if category_type == 'class':
                    # Try to find the class even with slight naming differences
                    if filter_value and filter_value.lower() != 'all':
                        # Case-insensitive lookup
                        target_class = next((k for k in generated_timetable.keys() if k.lower() == filter_value.lower()), None)
                        classes = [target_class] if target_class else []
                    else:
                        classes = list(generated_timetable.keys())

                    for class_name in classes:
                        class_data = generated_timetable.get(class_name)
                        if not class_data: continue
                        
                        df = format_timetable_df(class_data, lecture_duration)
                        # Excel sheet name limit is 31 chars and no special chars
                        safe_sheet_name = "".join(x for x in class_name if x.isalnum() or x in " -_")[:31]
                        df.to_excel(writer, sheet_name=safe_sheet_name or "Timetable")
                        
                        # Apply styling
                        worksheet = writer.sheets[safe_sheet_name or "Timetable"]
                        
                        # Set column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if cell.value:
                                        lines = str(cell.value).split('\n')
                                        curr_max = max(len(line) for line in lines)
                                        if curr_max > max_length:
                                            max_length = curr_max
                                except:
                                    pass
                            worksheet.column_dimensions[column_letter].width = max_length + 5

                        # Colors
                        theory_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
                        practical_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                        
                        for row in worksheet.iter_rows(min_row=2, min_col=2):
                            for cell in row:
                                if cell.value and isinstance(cell.value, str):
                                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                    if "[Theory]" in cell.value:
                                        cell.fill = theory_fill
                                    elif "[Practical]" in cell.value:
                                        cell.fill = practical_fill
                        
                        sheet_added = True
                
                elif category_type == 'teacher':
                    teacher_schedule = filter_timetable_by(generated_timetable, filter_value, 'teacher')
                    if teacher_schedule:
                        df = format_timetable_df(teacher_schedule, lecture_duration)
                        sheet_name = filter_value[:31] if filter_value else "Teacher"
                        df.to_excel(writer, sheet_name=sheet_name)
                        worksheet = writer.sheets[sheet_name]
                        # Apply same styling logic
                        for column in worksheet.columns:
                            worksheet.column_dimensions[column[0].column_letter].width = 25
                        for row in worksheet.iter_rows(min_row=2, min_col=2):
                            for cell in row:
                                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                if cell.value and isinstance(cell.value, str):
                                    if "[Theory]" in cell.value: cell.fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
                                    elif "[Practical]" in cell.value: cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                        sheet_added = True
                
                elif category_type == 'classroom':
                    classroom_schedule = filter_timetable_by(generated_timetable, filter_value, 'classroom')
                    if classroom_schedule:
                        df = format_timetable_df(classroom_schedule, lecture_duration)
                        sheet_name = filter_value[:31] if filter_value else "Classroom"
                        df.to_excel(writer, sheet_name=sheet_name)
                        worksheet = writer.sheets[sheet_name]
                        # Apply same styling logic
                        for column in worksheet.columns:
                            worksheet.column_dimensions[column[0].column_letter].width = 25
                        for row in worksheet.iter_rows(min_row=2, min_col=2):
                            for cell in row:
                                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                if cell.value and isinstance(cell.value, str):
                                    if "[Theory]" in cell.value: cell.fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
                                    elif "[Practical]" in cell.value: cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                        sheet_added = True
                
                # Fallback if no data was found to prevent Excel crash
                if not sheet_added:
                    df_empty = pd.DataFrame([["No data available for the selected filter"]], columns=["Notice"])
                    df_empty.to_excel(writer, sheet_name="No Data")
            
            output.seek(0)
            return send_file(output, as_attachment=True, download_name=f"timetable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        else:
            return jsonify({'error': 'Invalid export type'}), 400

        # For non-timetable exports
        if df.empty:
            df = pd.DataFrame(columns=['No data found'])
            
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=filename)

    except Exception as e:
        app.logger.error(f"Export error: {e}")
        return jsonify({'error': str(e)}), 500

def format_timetable_df(schedule_data, lecture_duration_mins):
    """Formats raw timetable JSON into a Pandas DataFrame grid."""
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    all_slots = set()
    for day in days:
        if day in schedule_data:
            all_slots.update(schedule_data[day].keys())
    
    sorted_slots = sorted(list(all_slots), key=lambda x: parse_time_robustly(x.split(' to ')[0]))
    
    grid = []
    total_lectures = 0
    for slot in sorted_slots:
        row = {'Time': slot}
        for day in days:
            cell = schedule_data.get(day, {}).get(slot, '-')
            if isinstance(cell, dict):
                total_lectures += 1
                s_type = cell.get('type', 'Theory')
                row[day] = f"{cell['subject']}\n({cell.get('classroom', cell.get('className', ''))})\n[{s_type}]"
            else:
                row[day] = cell
        grid.append(row)
    
    df = pd.DataFrame(grid)
    df.set_index('Time', inplace=True)
    
    # Calculate weekly hours
    total_hours = (total_lectures * lecture_duration_mins) / 60.0
    
    # Append total hours row
    df.loc['Total Weekly Hours'] = ['' for _ in range(len(days))]
    df.iloc[-1, 0] = f"{total_hours:.2f} hrs"
    
    return df

def filter_timetable_by(generated_timetable, value, target_type):
    """Filters a class-based timetable by teacher or classroom."""
    filtered = {}
    for class_name, days in generated_timetable.items():
        for day, slots in days.items():
            for slot, data in slots.items():
                match = False
                if target_type == 'teacher':
                    if value == 'all' or value in data['subject']:
                        match = True
                elif target_type == 'classroom':
                    if value == 'all' or value == data['classroom']:
                        match = True
                
                if match:
                    if day not in filtered: filtered[day] = {}
                    filtered[day][slot] = {**data, 'className': class_name}
    return filtered


# Staff Account Management
@app.route("/api/staff", methods=["GET"])
@login_required
@roles_required("admin")
def get_staff_accounts():
    try:
        with mysql_connection() as (mysql, cur):
            cur.execute("""SELECT user_id, name, email, role FROM users WHERE parent_admin_id = %s""", (session.get("user_id"),))
            staff = cur.fetchall()
            return jsonify(staff)
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/staff", methods=["POST"])
@login_required
@roles_required("admin")
def create_staff_account():
    data = request.json
    name, email, password = data.get("name"), data.get("email"), data.get("password")
    if not all([name, email, password]): return jsonify({"error": "Missing fields"}), 400
    try:
        hashed_password = generate_password_hash(password)
        with mysql_connection() as (mysql, cur):
            cur.execute("SELECT user_id FROM users WHERE email = %s", (email,))
            if cur.fetchone(): return jsonify({"error": "Email exists"}), 409
            cur.execute("""INSERT INTO users (name, email, password_hash, role, parent_admin_id, created_at) VALUES (%s, %s, %s, "teacher", %s, %s)""", (name, email, hashed_password, session.get("user_id"), datetime.now()))
            mysql.commit()
            return jsonify({"message": "Success"}), 201
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/staff/<int:staff_user_id>", methods=["DELETE"])
@login_required
@roles_required("admin")
def delete_staff_account(staff_user_id):
    try:
        with mysql_connection() as (mysql, cur):
            cur.execute("DELETE FROM users WHERE user_id = %s AND parent_admin_id = %s", (staff_user_id, session.get("user_id")))
            mysql.commit()
            return jsonify({"message": "Deleted"})
    except Exception as e: return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True)
